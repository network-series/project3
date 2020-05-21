# !/usr/bin/python
# -*- coding: utf-8 -*-
# @time    : 2020/5/13
# @author  : shyanne shan
# @function: 软件接口

from pydantic import BaseModel
from fastapi import FastAPI
import random
import openpyxl
def made():  # 生成激活码
    activation_code = ''.join(random.sample('0123456789', 10))
    return activation_code
def judge(a):  # 判断生成的激活码是否和字典中存在的激活码重复
    new_made = made()
    for k in a:
        if a[k] != new_made:
            return new_made
        else:
            judge()

def generateLic(number):
    a = {1: made()}
#usrId为购买许可证的用户名，licID为许可证编号，licen为许可证
    for i in range(2, number + 1):
        a[i] = judge(a)
    sorted_list = sorted(a.items(), key=lambda x: x[0], reverse=False)
    print(sorted_list)
    return a


class Item(BaseModel):
    usrID: str
    password: str


app = FastAPI()
@app.post("/login")
async def create_item(item: Item):
    print(item.usrID)
    usrID=item.usrID
    password=item.password
    wb = openpyxl.load_workbook('D:\pro\project3\Server-Client\information.xlsx')
    ws = wb['Sheet1']
    row = ws.max_row
    valid = False
    for i in range(2, row + 1):
        if (usrID == ws.cell(row=i, column=2).value):
            if (password == ws.cell(row=i, column=3).value):  # ID和密码都正确 进行对应赋值
                valid = True
                data = ws.cell(row=i, column=4).value
                num = ws.cell(row=i, column=6).value
    if (valid):
        res = {
            "code": 1,  #
            "usrID": usrID,
            "password":password,
            "data": data,
            "num": num
        }
    else:
        res = {
            "code": 5  # 非法登陆
        }
    return res

class Item(BaseModel):
    usrID: str = None
    password: str = None
@app.post('/register')
def calculate(request_data: Item):
    new_usrID = request_data.usrID
    new_password=request_data.password
    print(request_data)
    #插入数据库
    wb = openpyxl.load_workbook('D:\pro\project3\Server-Client\information.xlsx')
    ws = wb['Sheet1']
    row = ws.max_row
    ws.cell(row=row + 1, column=2).value = new_usrID
    ws.cell(row=row + 1, column=3).value = new_password
    wb.save('D:\pro\project3\Server-Client\information.xlsx')
    res = {
            "code": 1  #
     }
    return res

class Item(BaseModel):
    usrID: str
    type: int
@app.post('/buy')
def calculate(request_data: Item):
    result = generateLic(1)  # 此处生成许可证后还需写入数据库
    print(result[1])
    print(request_data)
    data = result[1]
    if type == 0:
        num = 10
    else:
        num = 50
    wb = openpyxl.load_workbook('D:\pro\project3\Server-Client\information.xlsx')
    ws = wb['Sheet1']
    row = ws.max_row
    for i in range(2, row + 1):
        if (request_data.usrID == ws.cell(row=i, column=2).value):
            ws.cell(row=i, column=4).value = data  # 许可证内容写入Excel
            ws.cell(row=i, column=5).value = num
            wb.save('D:\pro\project3\Server-Client\information.xlsx')
    res = {"code": 1,  # 表示获取成功
           "usrID": request_data.usrID,
           "data": data,  # 许可证列表
           "num": num  # 该账号剩余许可证数量
           }
    return res


if __name__ == '__main__':
    import uvicorn

    uvicorn.run(app=app,
                host="127.0.0.1",
                port=8000,
                workers=1)